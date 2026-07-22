import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "venues.csv"
COPY_SOURCE = ROOT / "data" / "venue-display-copy.csv"
OUTPUT = ROOT / "今晚去哪-首批夜店资料.xlsx"

INK = "16181D"
WHITE = "FFFFFF"
MUTED = "68707D"
GRID = "D8DCE3"
LIGHT = "F5F6F8"
SHANGHAI = "FFE3DB"
HANGZHOU = "DDF2EC"
INS = "E4E8FF"
WARNING = "FFF0C7"
CORAL = "F26B4E"
TEAL = "168A7A"


def set_cell(cell, value, *, bold=False, color=INK, size=11, fill=None):
    cell.value = value
    cell.font = Font(name="Arial Unicode MS", size=size, bold=bold, color=color)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def style_title(ws, title, subtitle, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    set_cell(ws.cell(1, 1), title, bold=True, color=WHITE, size=18, fill=INK)
    set_cell(ws.cell(2, 1), subtitle, color="D7DBE2", size=10, fill=INK)
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 28


def style_table(ws, header_row, widths):
    thin = Side(style="thin", color=GRID)
    for cell in ws[header_row]:
        cell.font = Font(name="Arial Unicode MS", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[header_row].height = 30
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.sheet_view.showGridLines = False


def status_text(confidence):
    return {
        "高": "资料较完整",
        "中高": "部分待确认",
        "中": "需要亲测",
        "中低": "暂缓上线",
    }.get(confidence, "待整理")


def display_name(row):
    if row["type"] == "space":
        return f"INS · {row['name']}"
    return row["name"]


with SOURCE.open(encoding="utf-8-sig", newline="") as source:
    venues = list(csv.DictReader(source))

with COPY_SOURCE.open(encoding="utf-8-sig", newline="") as source:
    display_copy = {row["id"]: row for row in csv.DictReader(source)}

assert {venue["id"] for venue in venues} == set(display_copy)

wb = Workbook()
ws = wb.active
ws.title = "快速浏览"
style_title(
    ws,
    "今晚去哪｜首批夜店资料",
    "上海 + 杭州 · 这是公开资料底稿，先看定位是否准确，再由真实体验校正。",
    9,
)

headers = ["城市", "夜店", "一句话定位", "Club标签", "适合标签", "雷点标签", "主要音乐", "第一次去注意", "资料状态"]
for col, header in enumerate(headers, start=1):
    set_cell(ws.cell(4, col), header)

style_table(ws, 4, [10, 19, 30, 28, 24, 24, 27, 29, 14])

for row_index, venue in enumerate(venues, start=5):
    copy = display_copy[venue["id"]]
    city_label = "上海 · INS" if venue["type"] == "space" else venue["city"]
    values = [
        city_label,
        display_name(venue),
        copy["tagline"],
        copy["tags"],
        copy["suitable_tags"],
        copy["risk_tags"],
        venue["music"],
        venue["first_visit_tip"],
        status_text(venue["confidence"]),
    ]
    base_fill = INS if venue["type"] == "space" else (SHANGHAI if venue["city"] == "上海" else HANGZHOU)
    for col, value in enumerate(values, start=1):
        fill = base_fill if col <= 2 else (LIGHT if row_index % 2 else WHITE)
        if col == 9 and venue["confidence"] in {"中", "中低"}:
            fill = WARNING
        set_cell(ws.cell(row_index, col), value, bold=(col == 2), fill=fill)
        ws.cell(row_index, col).border = Border(bottom=Side(style="thin", color=GRID))
    ws.row_dimensions[row_index].height = 72

ws.freeze_panes = "C5"
ws.auto_filter.ref = f"A4:I{4 + len(venues)}"
ws.auto_filter.add_filter_column(0, ["上海", "上海 · INS", "杭州"])
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_title_rows = "1:4"


ins_ws = wb.create_sheet("INS楼层")
style_title(
    ins_ws,
    "INS 到底该去几楼？",
    "把 INS 当成一个夜生活商场：先按音乐和玩法选空间，不要进去以后再乱逛。",
    8,
)
ins_headers = ["空间", "楼层", "一句话定位", "Club标签", "适合标签", "雷点标签", "主要音乐", "第一次去注意"]
for col, header in enumerate(ins_headers, start=1):
    set_cell(ins_ws.cell(4, col), header)
style_table(ins_ws, 4, [17, 9, 29, 28, 23, 23, 28, 30])

floors = {
    "HUSH": "2F",
    "RADi": "3F",
    "Culture Club": "4F",
    "FRiENDS": "待确认",
    "La Fin": "6F",
    "OASIS": "B1",
}
ins_venues = [venue for venue in venues if venue["type"] == "space"]
for row_index, venue in enumerate(ins_venues, start=5):
    copy = display_copy[venue["id"]]
    values = [
        venue["name"],
        floors.get(venue["name"], "待确认"),
        copy["tagline"],
        copy["tags"],
        copy["suitable_tags"],
        copy["risk_tags"],
        venue["music"],
        venue["first_visit_tip"],
    ]
    for col, value in enumerate(values, start=1):
        fill = INS if col <= 2 else (LIGHT if row_index % 2 else WHITE)
        set_cell(ins_ws.cell(row_index, col), value, bold=(col == 1), fill=fill)
        ins_ws.cell(row_index, col).border = Border(bottom=Side(style="thin", color=GRID))
    ins_ws.row_dimensions[row_index].height = 74
ins_ws.freeze_panes = "C5"
ins_ws.auto_filter.ref = f"A4:H{4 + len(ins_venues)}"
ins_ws.sheet_properties.pageSetUpPr.fitToPage = True
ins_ws.page_setup.fitToWidth = 1
ins_ws.page_setup.fitToHeight = 0
ins_ws.print_title_rows = "1:4"


questions_ws = wb.create_sheet("待你确认")
style_title(
    questions_ws,
    "最需要你补充的 8 件事",
    "不用一次填完。你亲自去过哪家，就先写哪家；真实感受比满分评价更有价值。",
    5,
)
question_headers = ["夜店", "需要确认的问题", "你的真实体验", "可以公开吗", "处理状态"]
for col, header in enumerate(question_headers, start=1):
    set_cell(questions_ws.cell(4, col), header)
style_table(questions_ws, 4, [20, 46, 52, 16, 16])

questions = [
    ("ABYSS", "新场地的闷热和拥挤是否仍然明显？"),
    ("坎", "现在更像纯 Hip-hop 小厅，还是卡座社交场？"),
    ("ORII", "当前正式名称是 ORii Hub 还是 ORII CLUB？"),
    ("INS", "你去过的几个空间里，哪个最适合第一次去的人？为什么？"),
    ("INS", "Culture Club、FRiENDS 和 OASIS 的实际曲风分别是什么？"),
    ("NONOBOOM", "目前是否正常营业？它与 NONO 莱福是什么关系？"),
    ("NONO 莱福", "重装后和你去过的版本相比，变化有多大？"),
    ("404", "上海店当前地址和营业状态是什么？"),
]
for row_index, (venue, question) in enumerate(questions, start=5):
    values = [venue, question, "", "", "待补充"]
    for col, value in enumerate(values, start=1):
        fill = WARNING if col in {3, 4} else (LIGHT if row_index % 2 else WHITE)
        set_cell(questions_ws.cell(row_index, col), value, bold=(col == 1), fill=fill)
        questions_ws.cell(row_index, col).border = Border(bottom=Side(style="thin", color=GRID))
    questions_ws.row_dimensions[row_index].height = 58
questions_ws.freeze_panes = "C5"
questions_ws.auto_filter.ref = f"A4:E{4 + len(questions)}"
questions_ws.sheet_properties.pageSetUpPr.fitToPage = True
questions_ws.page_setup.fitToWidth = 1
questions_ws.page_setup.fitToHeight = 0
questions_ws.print_title_rows = "1:4"


for sheet in wb.worksheets:
    sheet.sheet_properties.tabColor = {"快速浏览": CORAL, "INS楼层": "6574D9", "待你确认": TEAL}[sheet.title]
    sheet.sheet_view.zoomScale = 85
    sheet.auto_filter.ref = sheet.auto_filter.ref

wb.save(OUTPUT)

# Reopen once so generation fails immediately if the workbook is malformed.
check = load_workbook(OUTPUT)
assert check.sheetnames == ["快速浏览", "INS楼层", "待你确认"]
assert check["快速浏览"].max_row == 4 + len(venues)
check.close()
print(OUTPUT)
